# ============================================================
# app.py  —  Flask Web Frontend for Smart Attendance System
# Wraps: takeImage.py, trainImage.py, automaticAttedance.py
#
# Adds:
#   - Login / Sign-up / Logout (local CSV-backed accounts)
#   - PER-USER DATA ISOLATION: every account gets its own private
#     folder under UserData/<username>/ for students, models,
#     attendance CSVs and phone-detection screenshots.
#   - Phone-detection screenshots gallery in Records
#   - Delete a subject (CSVs + screenshots) from Records
#   - Subject search + "Download as Excel" from Records
#
# Run:  python app.py   then open http://localhost:5000
# ============================================================
import os, csv, json, time, threading, datetime, glob, shutil, io, re
from functools import wraps
from flask import (Flask, render_template, request, jsonify, Response,
                   stream_with_context, redirect, url_for, session,
                   send_file, send_from_directory, flash, abort)
from werkzeug.security import generate_password_hash, check_password_hash
import cv2
import numpy as np
import pandas as pd
import pyttsx3
from openpyxl import Workbook

# ---- backend modules (your existing files) ----
import takeImage as ti
import trainImage as tri
import automaticAttedance as aa

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
USERS_DIR   = os.path.join(BASE_DIR, "Users")
USERS_FILE  = os.path.join(USERS_DIR, "users.csv")
USERDATA_ROOT = os.path.join(BASE_DIR, "UserData")   # parent of all per-user folders
os.makedirs(USERS_DIR, exist_ok=True)
os.makedirs(USERDATA_ROOT, exist_ok=True)

app = Flask(__name__)
# Use SESSION_SECRET from environment if available, otherwise fall back to a
# stable file-based secret so existing sessions survive restarts.
_secret_file = os.path.join(USERS_DIR, ".secret")
if os.environ.get("SESSION_SECRET"):
    app.secret_key = os.environ["SESSION_SECRET"]
else:
    if not os.path.exists(_secret_file):
        with open(_secret_file, "wb") as f:
            f.write(os.urandom(32))
    with open(_secret_file, "rb") as f:
        app.secret_key = f.read()

# ---- shared camera state (only one webcam exists, so only one user can
#       run registration / attendance at a time — protected by the lock). ----
camera_lock   = threading.Lock()
active_camera = None

# ---- TTS (thread-safe wrapper) ----
_tts_lock = threading.Lock()
def tts(text):
    def _run():
        with _tts_lock:
            try:
                e = pyttsx3.init(); e.say(text); e.runAndWait()
            except: pass
    threading.Thread(target=_run, daemon=True).start()


# ============================================================
# AUTH — simple local CSV-backed accounts
# ============================================================
_users_lock = threading.Lock()
_USERNAME_RE = re.compile(r"^[A-Za-z0-9_.-]{3,32}$")

def _safe_username(name):
    """Allow only letters/digits/_.- to keep the folder name safe."""
    if not name: return None
    if not _USERNAME_RE.match(name): return None
    if name in (".", ".."): return None
    return name

def _load_users():
    """Return dict: username -> {email, password_hash}."""
    users = {}
    if not os.path.exists(USERS_FILE):
        return users
    try:
        with open(USERS_FILE, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                u = (row.get("username") or "").strip()
                if u:
                    users[u] = {
                        "email": (row.get("email") or "").strip(),
                        "password_hash": row.get("password_hash") or "",
                    }
    except Exception as e:
        print(f"User file read error: {e}")
    return users

def _save_user(username, email, password_hash):
    write_header = not os.path.isfile(USERS_FILE) or os.path.getsize(USERS_FILE) == 0
    with open(USERS_FILE, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if write_header:
            w.writerow(["username", "email", "password_hash", "created_at"])
        w.writerow([username, email, password_hash,
                    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user"):
            if request.path.startswith("/api/") or request.path.startswith("/video/"):
                return jsonify({"ok": False, "msg": "Login required"}), 401
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)
    return wrapped

@app.context_processor
def inject_user():
    return {"current_user": session.get("user")}


# ============================================================
# PER-USER DATA ROOT  —  the core of data isolation
# ============================================================
def user_root():
    """Return the *current* logged-in user's private data folder, creating
    its subdirs on first use. Aborts with 401 if the session is missing.

    Layout:
        UserData/<username>/
            StudentDetails/studentdetails.csv
            TrainingImage/<id_name>/...
            TrainingImageLabel/Trainer_<id>.yml
            Attendance/<subject>/<csv files>
            phone_detections/<subject>/<jpg files>
    """
    u = session.get("user")
    if not u or not _safe_username(u):
        abort(401)
    base = os.path.join(USERDATA_ROOT, u)
    for sub in ("StudentDetails", "TrainingImage", "TrainingImageLabel",
                "Attendance", "phone_detections"):
        os.makedirs(os.path.join(base, sub), exist_ok=True)
    return base

# Convenience accessors (always read from current user)
def U_STUDENT_CSV():  return os.path.join(user_root(), "StudentDetails", "studentdetails.csv")
def U_TRAIN_IMG():    return os.path.join(user_root(), "TrainingImage")
def U_TRAIN_LABEL():  return os.path.join(user_root(), "TrainingImageLabel")
def U_ATTENDANCE():   return os.path.join(user_root(), "Attendance")
def U_PHONESHOTS():   return os.path.join(user_root(), "phone_detections")


# ============================================================
# AUTH ROUTES
# ============================================================
@app.route("/signup", methods=["GET", "POST"])
def signup():
    if session.get("user"):
        return redirect(url_for("index"))
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        email    = (request.form.get("email") or "").strip()
        pwd      = request.form.get("password") or ""
        confirm  = request.form.get("confirm") or ""
        err = None
        if not username or not email or not pwd:
            err = "All fields are required."
        elif not _safe_username(username):
            err = "Username must be 3–32 chars: letters, digits, _ . - only."
        elif len(pwd) < 6:
            err = "Password must be at least 6 characters."
        elif pwd != confirm:
            err = "Passwords do not match."
        else:
            with _users_lock:
                users = _load_users()
                if username in users:
                    err = "That username is already taken."
                else:
                    _save_user(username, email, generate_password_hash(pwd))
                    # Pre-create the user's private data folder
                    os.makedirs(os.path.join(USERDATA_ROOT, username), exist_ok=True)
        if err:
            return render_template("signup.html", error=err,
                                   username=username, email=email)
        # auto-login after sign-up
        session["user"] = username
        return redirect(url_for("index"))
    return render_template("signup.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("user"):
        return redirect(url_for("index"))
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        pwd      = request.form.get("password") or ""
        users = _load_users()
        u = users.get(username)
        if not u or not check_password_hash(u["password_hash"], pwd):
            return render_template("login.html",
                                   error="Invalid username or password.",
                                   username=username)
        session["user"] = username
        nxt = request.args.get("next") or url_for("index")
        return redirect(nxt)
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect(url_for("login"))


# ============================================================
# ROUTES — Pages
# ============================================================
@app.route("/")
@login_required
def index():
    return render_template("index.html")

@app.route("/register")
@login_required
def register():
    return render_template("register.html")

@app.route("/attendance")
@login_required
def attendance():
    return render_template("attendance.html")

@app.route("/records")
@login_required
def records():
    return render_template("records.html")

# ============================================================
# ROUTES — Registration API
# ============================================================
register_state = {
    "running": False, "status": "idle",
    "angle": "front", "counters": {},
    "sampleNum": 0, "done": False, "error": "", "owner": ""
}

def _run_registration(enrollment, name, user_dir):
    """Background worker. user_dir = UserData/<username>/ — captured BEFORE
    spawning this thread (since session is request-scoped).

    This thread is the SOLE owner of the webcam during registration. It both
    captures the face crops (saved to disk) AND publishes annotated preview
    JPEGs into _reg_frame_buf so /video/register can stream them to the
    browser without ever touching the camera itself. Opening the camera
    twice on Windows DirectShow causes the second open to fail silently,
    which used to leave cam.read() returning False forever and no images
    were ever captured.
    """
    global register_state, active_camera, _reg_frame_buf
    register_state.update({"running": True, "done": False, "error": "",
                            "status": "starting", "sampleNum": 0,
                            "counters": {a: 0 for a in ['front','left','right','up','down']},
                            "owner": os.path.basename(user_dir)})
    cam = None
    try:
        yunet_path = os.path.join(BASE_DIR, "face_detection_yunet_2023mar.onnx")
        train_path = os.path.join(user_dir, "TrainingImage")
        os.makedirs(train_path, exist_ok=True)

        yunet = ti.initialize_yunet(yunet_path)
        if yunet is None:
            register_state.update({"error":
                "YuNet model file not found (face_detection_yunet_2023mar.onnx). "
                "Make sure it is in the project folder.",
                "running": False, "status": "error"})
            return

        directory = f"{enrollment}_{name}"
        path = os.path.join(train_path, directory)
        os.makedirs(path, exist_ok=True)

        angle_seq     = ['front', 'left', 'right', 'up', 'down']
        target        = 500
        angle_idx     = 0
        current_angle = 'front'
        last_cap      = 0; cap_delay = 0.05; sample = 0
        guidance = {'front':"LOOK STRAIGHT", 'left':"TURN HEAD LEFT",
                    'right':"TURN HEAD RIGHT", 'up':"TILT HEAD UP",
                    'down':"TILT HEAD DOWN"}

        # ---- Open camera ONCE; this thread is the sole owner ----
        with camera_lock:
            # DirectShow first (best on Windows), then default backend.
            cam = cv2.VideoCapture(0, cv2.CAP_DSHOW)
            if not cam.isOpened():
                cam.release()
                cam = cv2.VideoCapture(0)
            if not cam.isOpened():
                register_state.update({"error":
                    "Cannot open webcam. Close any other app using the camera "
                    "(Zoom, Teams, browser tabs) and try again.",
                    "running": False, "status": "error"})
                return
            cam.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
            cam.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
            active_camera = cam

        # Warm up: many webcams return a few black frames after open
        for _ in range(15):
            ok, _f = cam.read()
            if ok: break
            time.sleep(0.05)

        register_state["status"] = "capturing"
        consecutive_read_fail = 0

        while True:
            if not register_state["running"]:
                break
            ret, img = cam.read()
            if not ret or img is None:
                consecutive_read_fail += 1
                if consecutive_read_fail > 100:
                    register_state["error"] = ("Camera stopped returning frames. "
                        "Unplug and re-plug the webcam, then try again.")
                    break
                time.sleep(0.02)
                continue
            consecutive_read_fail = 0

            img = cv2.flip(img, 1)
            clean = img.copy()                 # used for SAVING (no overlay)
            det = ti.detect_face_yunet(img, yunet)
            register_state["angle"] = current_angle

            if det:
                x, y, w, h, conf = det
                now = time.time()
                if now - last_cap >= cap_delay and register_state["counters"][current_angle] < target:
                    margin = int(min(w, h) * 0.1)
                    y1 = max(0, y - margin); y2 = min(clean.shape[0], y + h + margin)
                    x1 = max(0, x - margin); x2 = min(clean.shape[1], x + w + margin)
                    face = cv2.cvtColor(clean[y1:y2, x1:x2], cv2.COLOR_BGR2GRAY)
                    face = cv2.resize(face, (200, 200))
                    sample += 1
                    register_state["counters"][current_angle] += 1
                    register_state["sampleNum"] = sample
                    fname = f"{name}_{enrollment}_{current_angle}_{register_state['counters'][current_angle]}.jpg"
                    cv2.imwrite(os.path.join(path, fname), face)
                    last_cap = now

                # Draw overlay on PREVIEW frame only (clean copy already saved)
                cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 120), 2)
                cv2.putText(img, f"Face {conf:.2f}", (x, max(0, y - 8)),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.55, (0, 255, 120), 2)
            else:
                cv2.putText(img, "No face detected — move closer / better light",
                            (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 100, 255), 2)

            # ---- Build the preview overlay ----
            cv2.putText(img, guidance.get(current_angle, ""),
                        (10, img.shape[0] - 20), cv2.FONT_HERSHEY_SIMPLEX,
                        0.85, (0, 220, 255), 2)
            tot = sum(register_state["counters"].values())
            cv2.putText(img,
                        f"{current_angle.upper()}  "
                        f"{register_state['counters'][current_angle]}/{target}   "
                        f"TOTAL {tot}/{target*5}",
                        (10, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 0), 2)

            # ---- Publish to MJPEG buffer for /video/register ----
            ok, jpg = cv2.imencode('.jpg', img, [cv2.IMWRITE_JPEG_QUALITY, 75])
            if ok:
                with _reg_frame_lock:
                    _reg_frame_buf = jpg.tobytes()

            # Advance angle when one is filled
            if register_state["counters"].get(current_angle, 0) >= target:
                if angle_idx < len(angle_seq) - 1:
                    angle_idx += 1
                    current_angle = angle_seq[angle_idx]

            if all(register_state["counters"].get(a, 0) >= target for a in angle_seq):
                break

        # Save to PER-USER CSV (only if we actually captured something)
        if sample > 0:
            csv_path = os.path.join(user_dir, "StudentDetails", "studentdetails.csv")
            os.makedirs(os.path.dirname(csv_path), exist_ok=True)
            write_header = not os.path.isfile(csv_path) or os.path.getsize(csv_path) == 0
            with open(csv_path, "a", newline="") as f:
                w = csv.writer(f)
                if write_header: w.writerow(["Enrollment", "Name"])
                w.writerow([enrollment, name])

        register_state.update({"running": False, "done": True, "status": "done",
                                "sampleNum": sample})
        if sample > 0:
            tts(f"Registration complete. {sample} images saved for {name}.")
    except Exception as e:
        import traceback; traceback.print_exc()
        register_state.update({"running": False, "error": str(e), "status": "error"})
    finally:
        if cam is not None:
            try: cam.release()
            except: pass
        # Give Windows DirectShow time to fully reset the device.
        # Without this, attendance started right after registration gets
        # blank frames because the driver is still tearing down internally.
        time.sleep(2.0)
        with camera_lock: active_camera = None
        # Clear the preview buffer so the browser stops showing the last frame
        with _reg_frame_lock:
            _reg_frame_buf = None


@app.route("/api/register/start", methods=["POST"])
@login_required
def api_register_start():
    if register_state["running"]:
        return jsonify({"ok": False, "msg": "Registration already running"})
    data = request.json
    enrollment = data.get("enrollment", "").strip()
    name       = data.get("name", "").strip()
    if not enrollment or not name:
        return jsonify({"ok": False, "msg": "Enrollment and name required"})
    user_dir = user_root()   # capture BEFORE spawning thread
    threading.Thread(target=_run_registration,
                     args=(enrollment, name, user_dir), daemon=True).start()
    return jsonify({"ok": True})

@app.route("/api/register/status")
@login_required
def api_register_status():
    return jsonify(register_state)

@app.route("/api/register/stop", methods=["POST"])
@login_required
def api_register_stop():
    register_state["running"] = False
    return jsonify({"ok": True})

# ============================================================
# ROUTES — Training API
# ============================================================
train_state = {"running": False, "done": False, "msg": "", "error": "", "owner": ""}

def _run_training(user_dir):
    global train_state
    train_state = {"running": True, "done": False,
                   "msg": "Training in progress...", "error": "",
                   "owner": os.path.basename(user_dir)}
    try:
        class FakeLabel:
            def configure(self, **kw): train_state["msg"] = kw.get("text","")
        tri.TrainImage(
            haarcasecade_path="",
            trainimage_path=os.path.join(user_dir, "TrainingImage"),
            trainimagelabel_path=os.path.join(user_dir, "TrainingImageLabel", "Trainner.yml"),
            message=FakeLabel(),
            text_to_speech=tts
        )
        train_state.update({"running": False, "done": True})
    except Exception as e:
        train_state.update({"running": False, "error": str(e), "done": True})

@app.route("/api/train/start", methods=["POST"])
@login_required
def api_train_start():
    if train_state["running"]:
        return jsonify({"ok": False, "msg": "Training already running"})
    user_dir = user_root()
    threading.Thread(target=_run_training, args=(user_dir,), daemon=True).start()
    return jsonify({"ok": True})

@app.route("/api/train/status")
@login_required
def api_train_status():
    return jsonify(train_state)


# ============================================================
# ROUTES — Attendance API + Live Camera Stream
# ============================================================
attend_state = {
    "running": False, "slot_idx": 0, "phase": "idle",
    "slot": "", "remaining": 0, "present": {}, "phone": False,
    "slots": [], "subject": "", "done": False, "error": "",
    "face_labels": [], "owner": ""
}
_frame_buf = None   # latest annotated JPEG bytes (attendance stream)
_frame_lock = threading.Lock()

# ---- Registration stream buffer ----
# CRITICAL: On Windows DirectShow only ONE process can hold the webcam at a
# time. Previously the registration worker thread opened the camera AND the
# /video/register MJPEG endpoint also tried to open it -> the second open
# either failed or invalidated the first handle, so cam.read() returned
# False forever and no images were captured. Now the worker thread is the
# SOLE owner of the camera and pushes annotated JPEG bytes into this buffer;
# the MJPEG endpoint just yields whatever is in the buffer.
_reg_frame_buf  = None
_reg_frame_lock = threading.Lock()

def _run_attendance(subject, slots, user_dir):
    global attend_state, _frame_buf, active_camera
    attend_state.update({"running": True, "done": False, "error": "", "subject": subject,
                         "slots": slots, "present": {}, "phase": "starting",
                         "owner": os.path.basename(user_dir)})

    # Point automaticAttedance.py at THIS user's data BEFORE using it.
    aa.set_user_root(user_dir)

    recognizers = aa.load_per_student_recognizers()
    if not recognizers:
        attend_state.update({"running": False, "error": "No trained models found. Train first."}); return

    # Load ArcFace embeddings — these are the primary recognizer.
    # They handle bearded/similar-looking students correctly where LBPH fails.
    embeddings = aa.load_per_student_embeddings()

    try:
        student_df = pd.read_csv(aa.STUDENTDETAIL_PATH)
    except Exception as e:
        attend_state.update({"running": False, "error": f"Cannot read student CSV: {e}"}); return

    cam = cv2.VideoCapture(0, cv2.CAP_DSHOW)
    if not cam.isOpened():
        cam.release(); cam = cv2.VideoCapture(0)
    # Use 640x480 — same as registration — so DirectShow reuses the existing
    # filter graph instead of rebuilding it. Rebuilding right after registration
    # released the camera causes a hang on Windows.
    cam.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
    cam.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
    if not cam.isOpened():
        attend_state.update({"running": False, "error": "Cannot open camera"}); return

    # Drain the initial blank frames DirectShow emits after a fresh open,
    # especially right after registration released the camera.
    for _ in range(30):
        ok, _warmup = cam.read()
        if ok:
            break
        time.sleep(0.05)

    with camera_lock: active_camera = cam

    try:
        for slot_idx, slot in enumerate(slots):
            # ---- WAIT PHASE ----
            attend_state["phase"] = "waiting"; attend_state["slot"] = slot
            attend_state["slot_idx"] = slot_idx
            wait_secs  = aa.wait_until(slot)
            start_wait = time.time()

            while time.time() - start_wait < wait_secs:
                if not attend_state["running"]: break
                ret, frame = cam.read()
                if not ret: continue
                frame = cv2.flip(frame, 1)
                face_boxes = aa.detect_faces_insightface(frame)
                frame, phone_ok = aa.detect_phone(frame, face_boxes, subject)
                rem = int(wait_secs - (time.time() - start_wait))
                attend_state["remaining"] = rem; attend_state["phone"] = phone_ok
                cv2.putText(frame, f"Waiting for slot {slot_idx+1} at {slot} | {rem}s",
                    (10,35), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0,220,255), 2)
                if phone_ok:
                    cv2.putText(frame,"PHONE DETECTED!",(20,85),
                        cv2.FONT_HERSHEY_DUPLEX,1.4,(0,0,255),3)
                _, jpg = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 70])
                with _frame_lock: _frame_buf = jpg.tobytes()

            # ---- ATTENDANCE PHASE ----
            attend_state["phase"] = "marking"
            att_start = time.time(); ATT_DUR = 60
            slot_attendance = {}

            while time.time() - att_start < ATT_DUR:
                if not attend_state["running"]: break
                ret, frame = cam.read()
                if not ret: continue
                frame = cv2.flip(frame, 1)
                gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

                # One InsightFace pass → get both face boxes AND ArcFace embeddings
                faces_with_emb = aa.get_faces_with_embeddings(frame)
                face_boxes = [box for box, _ in faces_with_emb]
                labels = []

                for (fx, fy, fw, fh), emb in faces_with_emb:
                    fg = gray[fy:fy+fh, fx:fx+fw]
                    if fg.size == 0: continue

                    # Try ArcFace embedding recognition first
                    eid = None; name = "Unknown"; label_conf = ""
                    if embeddings and emb is not None:
                        eid, name, sim = aa.recognize_face_embedding(emb, embeddings, student_df)
                        if eid is not None:
                            label_conf = f"sim={sim:.2f}"
                        else:
                            eid, name, dist = aa.recognize_face(fg, recognizers, student_df)
                            label_conf = f"d={dist:.0f}"
                    else:
                        eid, name, dist = aa.recognize_face(fg, recognizers, student_df)
                        label_conf = f"d={dist:.0f}"

                    if eid:
                        label = f"{name} ({eid}) {label_conf}"; color = (0,255,0)
                        slot_attendance[eid] = name
                        attend_state["present"][str(eid)] = name
                    else:
                        label = f"Unknown {label_conf}"; color = (0,0,255)
                    labels.append(label)
                    cv2.rectangle(frame,(fx,fy),(fx+fw,fy+fh),color,2)
                    cv2.putText(frame,label,(fx,fy-10),cv2.FONT_HERSHEY_SIMPLEX,0.65,color,2)

                frame, phone_ok = aa.detect_phone(frame, face_boxes, subject)
                rem = int(ATT_DUR - (time.time() - att_start))
                attend_state.update({"remaining": rem, "phone": phone_ok, "face_labels": labels})
                cv2.putText(frame,f"Slot {slot_idx+1} | {slot} | {rem}s | Present:{len(slot_attendance)}",
                    (10,35),cv2.FONT_HERSHEY_SIMPLEX,0.75,(0,220,255),2)
                if phone_ok:
                    cv2.putText(frame,"PHONE DETECTED!",(20,85),cv2.FONT_HERSHEY_DUPLEX,1.4,(0,0,255),3)
                _, jpg = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 70])
                with _frame_lock: _frame_buf = jpg.tobytes()

            # ---- SAVE (per-user folder) ----
            if slot_attendance:
                date    = datetime.datetime.now().strftime("%Y-%m-%d")
                out_dir = os.path.join(user_dir, "Attendance", subject)
                os.makedirs(out_dir, exist_ok=True)
                fname   = f"{subject}_{date}_{slot.replace(':','-')}.csv"
                pd.DataFrame([(k,v) for k,v in slot_attendance.items()],
                    columns=["Enrollment","Name"]).to_csv(os.path.join(out_dir,fname), index=False)

    except Exception as e:
        attend_state["error"] = str(e)
    finally:
        cam.release()
        with camera_lock: active_camera = None
        attend_state.update({"running": False, "done": True, "phase": "done"})


@app.route("/api/attendance/start", methods=["POST"])
@login_required
def api_attendance_start():
    if attend_state["running"]:
        return jsonify({"ok": False, "msg": "Attendance already running"})
    data    = request.json
    subject = data.get("subject","").strip()
    slots   = data.get("slots", [])
    if not subject: return jsonify({"ok": False, "msg": "Subject required"})
    if len(slots) != 3: return jsonify({"ok": False, "msg": "Exactly 3 slots required"})
    user_dir = user_root()
    threading.Thread(target=_run_attendance,
                     args=(subject, slots, user_dir), daemon=True).start()
    return jsonify({"ok": True})

@app.route("/api/attendance/status")
@login_required
def api_attendance_status():
    return jsonify(attend_state)

@app.route("/api/attendance/stop", methods=["POST"])
@login_required
def api_attendance_stop():
    attend_state["running"] = False
    return jsonify({"ok": True})

def _gen_register_stream():
    """MJPEG stream for the Register page.

    IMPORTANT: this generator does NOT open the webcam. It only reads JPEG
    bytes that the registration worker thread (_run_registration) puts into
    _reg_frame_buf. Opening the camera here in addition to the worker thread
    causes a Windows DirectShow conflict that silently breaks capture.
    """
    boundary = b'--frame\r\nContent-Type: image/jpeg\r\n\r\n'
    # Wait briefly for the worker thread to publish the first frame.
    waited = 0.0
    while True:
        with _reg_frame_lock:
            buf = _reg_frame_buf
        if buf is None:
            # No frame yet. If registration is not running either, stop.
            if not register_state.get("running") and waited > 1.0:
                return
            time.sleep(0.05); waited += 0.05
            if waited > 10.0:           # camera never produced a frame
                return
            continue
        yield (boundary + buf + b'\r\n')
        # If the worker has stopped AND cleared the buffer, the next loop
        # iteration will see buf is None and exit cleanly.
        time.sleep(0.03)                # ~30 fps cap

def _gen_attendance_stream():
    while True:
        time.sleep(0.04)
        with _frame_lock:
            buf = _frame_buf
        if buf:
            yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + buf + b'\r\n')

@app.route("/video/register")
@login_required
def video_register():
    return Response(stream_with_context(_gen_register_stream()),
                    mimetype="multipart/x-mixed-replace; boundary=frame")

@app.route("/video/attendance")
@login_required
def video_attendance():
    return Response(stream_with_context(_gen_attendance_stream()),
                    mimetype="multipart/x-mixed-replace; boundary=frame")

# ============================================================
# ROUTES — Records API   (all read PER-USER folders)
# ============================================================
def _safe_subject(name):
    if not name: return None
    if "/" in name or "\\" in name or ".." in name or name.startswith("."):
        return None
    return name

@app.route("/api/records/subjects")
@login_required
def api_records_subjects():
    base      = user_root()
    att_dir   = os.path.join(base, "Attendance")
    phone_dir = os.path.join(base, "phone_detections")
    subjects = set()
    if os.path.exists(att_dir):
        for d in os.listdir(att_dir):
            if os.path.isdir(os.path.join(att_dir, d)): subjects.add(d)
    if os.path.exists(phone_dir):
        for d in os.listdir(phone_dir):
            if os.path.isdir(os.path.join(phone_dir, d)): subjects.add(d)
    return jsonify(sorted(subjects))

@app.route("/api/records/sessions")
@login_required
def api_records_sessions():
    subject = _safe_subject(request.args.get("subject",""))
    if not subject: return jsonify([])
    folder  = os.path.join(user_root(), "Attendance", subject)
    if not os.path.exists(folder): return jsonify([])
    files = sorted(glob.glob(os.path.join(folder, "*.csv")))
    return jsonify([os.path.basename(f) for f in files])

@app.route("/api/records/data")
@login_required
def api_records_data():
    subject = _safe_subject(request.args.get("subject",""))
    session_name = request.args.get("session","")
    if not subject or not session_name or "/" in session_name or "\\" in session_name:
        return jsonify({"headers":[],"rows":[]})
    fpath = os.path.join(user_root(), "Attendance", subject, session_name)
    if not os.path.exists(fpath): return jsonify({"headers":[],"rows":[]})
    df = pd.read_csv(fpath)
    return jsonify({"headers": list(df.columns), "rows": df.values.tolist()})

@app.route("/api/records/students")
@login_required
def api_records_students():
    try:
        df = pd.read_csv(os.path.join(user_root(), "StudentDetails", "studentdetails.csv"))
        return jsonify(df.to_dict(orient="records"))
    except:
        return jsonify([])

@app.route("/api/records/delete_student", methods=["POST"])
@login_required
def api_records_delete_student():
    data = request.get_json(silent=True) or {}
    try:
        enrollment = int(data.get("enrollment", 0))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "msg": "Invalid enrollment ID"}), 400
    if not enrollment:
        return jsonify({"ok": False, "msg": "Enrollment ID required"}), 400

    base         = user_root()
    csv_path     = os.path.join(base, "StudentDetails", "studentdetails.csv")
    img_dir      = os.path.join(base, "TrainingImage")
    label_dir    = os.path.join(base, "TrainingImageLabel")
    student_name = str(enrollment)

    # 1. Remove from StudentDetails CSV
    if os.path.exists(csv_path):
        try:
            df = pd.read_csv(csv_path)
            # Cast to int for safe comparison (CSV may store values as float)
            mask = df["Enrollment"].astype(int) == enrollment
            match = df[mask]
            if not match.empty:
                student_name = str(match["Name"].values[0])
            df = df[~mask]
            df.to_csv(csv_path, index=False)
        except Exception as e:
            return jsonify({"ok": False, "msg": f"CSV error: {e}"}), 500

    # 2. Remove training-image folder (e.g. TrainingImage/101_StudentName/)
    if os.path.isdir(img_dir):
        for folder in os.listdir(img_dir):
            try:
                if int(folder.split("_", 1)[0]) == enrollment:
                    shutil.rmtree(os.path.join(img_dir, folder), ignore_errors=True)
            except (ValueError, IndexError):
                pass

    # 3. Remove LBPH model file and ArcFace embedding file
    for fname in [f"Trainer_{enrollment}.yml", f"Embedding_{enrollment}.npy"]:
        fpath = os.path.join(label_dir, fname)
        if os.path.exists(fpath):
            try:
                os.remove(fpath)
            except Exception:
                pass

    return jsonify({"ok": True, "name": student_name})

@app.route("/api/records/download_students")
@login_required
def api_records_download_students():
    csv_path = os.path.join(user_root(), "StudentDetails", "studentdetails.csv")
    if not os.path.exists(csv_path):
        return jsonify({"ok": False, "msg": "No students registered"}), 404
    try:
        df = pd.read_csv(csv_path)
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=f"students_{today}.xlsx")

# ---------- Phone-detection screenshots gallery ----------
@app.route("/api/records/phone_screenshots")
@login_required
def api_records_phone_screenshots():
    subject = _safe_subject(request.args.get("subject",""))
    if not subject: return jsonify([])
    folder = os.path.join(user_root(), "phone_detections", subject)
    if not os.path.exists(folder): return jsonify([])
    out = []
    for f in sorted(glob.glob(os.path.join(folder, "*.*")), reverse=True):
        if not os.path.isfile(f): continue
        ext = os.path.splitext(f)[1].lower()
        if ext not in (".jpg", ".jpeg", ".png"): continue
        out.append({
            "filename": os.path.basename(f),
            "size": os.path.getsize(f),
            "mtime": int(os.path.getmtime(f)),
        })
    return jsonify(out)

@app.route("/phone_image/<subject>/<filename>")
@login_required
def phone_image(subject, filename):
    subject = _safe_subject(subject)
    if not subject or "/" in filename or "\\" in filename or ".." in filename:
        abort(404)
    folder = os.path.join(user_root(), "phone_detections", subject)
    if not os.path.exists(os.path.join(folder, filename)):
        abort(404)
    return send_from_directory(folder, filename)

# ---------- Delete a subject (CSVs + screenshots) ----------
@app.route("/api/records/delete_subject", methods=["POST"])
@login_required
def api_records_delete_subject():
    data = request.get_json(silent=True) or {}
    subject = _safe_subject((data.get("subject") or "").strip())
    if not subject:
        return jsonify({"ok": False, "msg": "Invalid subject name"}), 400
    base = user_root()
    removed = []
    for sub in ("Attendance", "phone_detections"):
        p = os.path.join(base, sub, subject)
        if os.path.isdir(p):
            try:
                shutil.rmtree(p)
                removed.append(sub)
            except Exception as e:
                return jsonify({"ok": False, "msg": f"Failed to delete {sub}: {e}"}), 500
    if not removed:
        return jsonify({"ok": False, "msg": "Nothing found for that subject"}), 404
    return jsonify({"ok": True, "removed": removed})

# ---------- Download all sessions of a subject as one .xlsx ----------
@app.route("/api/records/download")
@login_required
def api_records_download():
    subject = _safe_subject(request.args.get("subject",""))
    if not subject:
        return jsonify({"ok": False, "msg": "Invalid subject"}), 400
    folder = os.path.join(user_root(), "Attendance", subject)
    if not os.path.isdir(folder):
        return jsonify({"ok": False, "msg": "No records for that subject"}), 404

    files = sorted(glob.glob(os.path.join(folder, "*.csv")))
    if not files:
        return jsonify({"ok": False, "msg": "No CSV sessions found"}), 404

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    used_names = set()
    for fpath in files:
        try:
            df = pd.read_csv(fpath)
        except Exception:
            continue
        raw = os.path.splitext(os.path.basename(fpath))[0]
        if raw.lower().startswith(subject.lower() + "_"):
            raw = raw[len(subject) + 1:]
        for ch in r":\/?*[]":
            raw = raw.replace(ch, "-")
        name = raw[:31] or "Session"
        base_name, n = name, 2
        while name in used_names:
            suffix = f" ({n})"
            name = (base_name[:31 - len(suffix)]) + suffix
            n += 1
        used_names.add(name)

        ws = wb.create_sheet(title=name)
        headers = list(df.columns)
        ws.append(headers)
        for row in df.itertuples(index=False, name=None):
            ws.append(list(row))

    if not wb.worksheets:
        wb.create_sheet(title="Empty")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    download_name = f"{subject}_attendance_{today}.xlsx"
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=download_name)


# ============================================================
if __name__ == "__main__":
    print("\n" + "="*55)
    print("  Smart Attendance Web App")
    print("  Open browser: http://localhost:5000")
    print("="*55 + "\n")
    app.run(debug=False, threaded=True, host="0.0.0.0", port=5000)
